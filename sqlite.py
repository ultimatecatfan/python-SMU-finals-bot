import sqlite3
from openpyxl import workbook,load_workbook
book = load_workbook('finals 2024.xlsx')
sheet = book.active
print(sheet)
print(f"Number of rows: {sheet.max_row} Type: {type(sheet.max_row)}")
con = sqlite3.connect("bot.db")
cur = con.cursor()
cur.execute("CREATE TABLE IF NOT EXISTS finals (school, name, section,examdate)")
data = []
for i in range(1,sheet.max_row + 1):
    num = str(i)
    row = sheet[num]
    print(type(row))
    cur.execute("INSERT INTO finals VALUES (?,?,?,?)",(row[0].value,row[1].value,row[2].value,row[3].value))
        
res = cur.execute('SELECT * from finals')
print(res.fetchall())