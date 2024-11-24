from typing import Final 
from telegram import Update, ReplyKeyboardMarkup,ReplyKeyboardRemove
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler

TOKEN: Final = '7802531576:AAHXH78AbPk-mYz-WjB13eYaF33M9J_qp7U'
BOT_USERNAME: Final = '@SMUFinalsBot'

import sqlite3
from openpyxl import workbook,load_workbook
book = load_workbook('finals 2024.xlsx')
sheet = book.active
print(sheet)
print(f"Number of rows: {sheet.max_row} Type: {type(sheet.max_row)}")
con = sqlite3.connect("bot.db")
cur = con.cursor()
cur.execute("CREATE TABLE IF NOT EXISTS finals (school, name, section,examdate)")
data = []
for i in range(1,sheet.max_row + 1):
    num = str(i)
    sheet_row = sheet[num]
    print(type(sheet_row))
    cur.execute("INSERT INTO finals VALUES (?,?,?,?)",(sheet_row[0].value,sheet_row[1].value,str(int(sheet_row[2].value)),sheet_row[3].value))
        
res = cur.execute('SELECT * from finals')
print(res.fetchall())


SCHOOL, COURSE_TITLE, SECTION = range(3) 

async def helpcommand(update: Update, context: ContextTypes.DEFAULT_TYPE):
    print("Help function is running... ")
    await update.message.reply_text("To start conversation, use command /start. \n To receive help, use command /help. \n To find your finals dates, use /date.")
    
async def startcommand(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Hello! Thank you for chatting!")

async def datecommand(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [['SCIS','CIS','SOB','SOA','SOL','SOE']] #
    await update.message.reply_text(
        '<b>Hello There!\nLet\'s get some details about the final paper you\'re sworrying about.\nWhat school does the course fall under? (Give the initials please, e.g SCIS) </b>',
        parse_mode='HTML',
        reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        )

    return SCHOOL #transition to the school state 

async def school(update:Update, context:ContextTypes.DEFAULT_TYPE):
    print("In school handler..",flush=True)
    context.user_data['school'] = update.message.text 
    await update.message.reply_text("Thanks for indicating the school offering the final exam.")
    await update.message.reply_text("Next, what is the course title? Give the full name please.")
    return COURSE_TITLE 

async def course_title(update:Update, context:ContextTypes.DEFAULT_TYPE):
    print("In course title handler..",flush=True)
    context.user_data['course_title'] = update.message.text 
    await update.message.reply_text("Thanks for indicating your course title. Next, what is your section number? Give a number, please.")
    return SECTION

async def section(update:Update,context:ContextTypes.DEFAULT_TYPE):
    print("In section handler..",flush=True)
    context.user_data['section'] = update.message.text 
    print(f"User data: {context.user_data}")
    result = cur.execute("SELECT examdate FROM finals WHERE school = ? AND name = ? AND section = ?",(context.user_data['school'],context.user_data['course_title'],str(context.user_data['section'])))
    if not result:
        await update.message.reply_text("No finals found! Try again?")
    else:
        row = result.fetchone()
        await update.message.reply_text(f"Your finals date/time is: {row[0]}")
        await update.message.reply_text("Thank you for using SMU Finals Bot. Good luck!")
    return ConversationHandler.END



    



async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Thank you for chatting with us today!")
    return ConversationHandler.END

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        msg_type = str(update.message.chat.type)  # Ensure msg_type is a string
        text = str(update.message.text)  # Ensure text is a string
        print(f"Message type: {msg_type} Received text: {text}")  # Debugging print statement
    except AttributeError as e:
        print(f"AttributeError: {e}")
        print(f"Update: {update}")
        return
    except Exception as e:
        print(f"Unexpected error: {e}")
        return


if __name__ == '__main__':

    try:
        app = Application.builder().token(TOKEN).build()
        print(f"Application built: {app}")  
        app.add_handler(CommandHandler('start', startcommand))
        app.add_handler(CommandHandler('help',helpcommand))
        Conversation_Handler = ConversationHandler(entry_points=[CommandHandler("date",datecommand)],
        states = {
            SCHOOL: [MessageHandler(filters.Regex("^(SCIS|CIS|SOB|SOA|SOL|SOE)$"),school)],
            COURSE_TITLE: [MessageHandler(filters.TEXT,course_title)], 
            SECTION: [MessageHandler(filters.TEXT,section)]

        },
        fallbacks= [CommandHandler('cancel',cancel)])
        app.add_handler(Conversation_Handler)
        app.add_handler(CommandHandler('cancel',cancel))
        app.run_polling(poll_interval=3)
    except Exception as e:
        print(f"Error in main: {e}")

    

